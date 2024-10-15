import os
import json
import logging
from openpyxl import Workbook


def get_database(database, db_path):
    db = [cont.__dict__ for cont in database]
    json_path = os.path.join(db_path, "db.json")
    with open(json_path, "w+") as f:
        json.dump(db, f, indent=4)
    keys = []
    wb = Workbook()
    ws = wb.active
    for i in range(len(db)):
        sub_obj = db[i]
        if i == 0:
            keys = list(sub_obj.keys())
            for k in range(len(keys)):
                ws.cell(row=(i + 1), column=(k + 1), value=keys[k])
        for j in range(len(keys)):
            ws.cell(row=(i + 2), column=(j + 1), value=sub_obj[keys[j]])
    wb.save(db_path + os.sep + "db.xlsx")
    logging.info("Containers Database Excel Created")
